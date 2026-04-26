"""
Fill missing zip_code in usda-verified.csv by matching license_usda
against the FOIA Excel file's License Number column.
Pulls zip from 'Shipping State/Province' column in the Excel file.
Only fills where zip_code is currently blank.
"""

import csv
import os

try:
    import openpyxl
except ImportError:
    print("Need openpyxl: pip install openpyxl")
    exit(1)

BASE    = r"c:\Users\uthin\Desktop\NEW-SEEDS-LAW\pipeline\04_human_in_the_loop"
XLSX    = os.path.join(BASE, "input", "2026-AMS-00094-F Final Response Records_Redacted.xlsx")
VERIFIED = os.path.join(BASE, "output", "usda-verified.csv")

# Load Excel
wb = openpyxl.load_workbook(XLSX, read_only=True)
ws = wb.active

# Find column indices
header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
headers = {cell.value.strip() if cell.value else "": cell.column - 1 for cell in header_row}

print(f"Excel headers: {list(headers.keys())}")

# Build lookup: license_number -> shipping_state_province
license_col = None
zip_col = None

for name, idx in headers.items():
    if "license" in name.lower() and "number" in name.lower():
        license_col = idx
    if "shipping" in name.lower():
        zip_col = idx

if license_col is None or zip_col is None:
    print(f"Could not find columns. License col: {license_col}, Zip col: {zip_col}")
    print("Available headers:")
    for name in headers:
        print(f"  '{name}'")
    exit(1)

print(f"Using: license col={license_col}, zip col={zip_col}")

xlsx_lookup = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    lic = str(row[license_col]).strip() if row[license_col] else ""
    zp  = str(row[zip_col]).strip() if row[zip_col] else ""
    if lic and zp:
        xlsx_lookup[lic] = zp

print(f"Excel lookup entries: {len(xlsx_lookup)}")
wb.close()

# Load verified CSV
with open(VERIFIED, newline="", encoding="utf-8-sig", errors="replace") as f:
    reader = csv.DictReader(f)
    fieldnames = reader.fieldnames
    rows = list(reader)

print(f"Verified CSV rows: {len(rows)}")

# Fill missing zip codes
filled = 0
not_found = 0
already_has = 0

for row in rows:
    if row.get("zip_code", "").strip():
        already_has += 1
        continue

    lic = row.get("license_usda", "").strip()
    if lic in xlsx_lookup:
        row["zip_code"] = xlsx_lookup[lic]
        filled += 1
    else:
        not_found += 1

print(f"\nResults:")
print(f"  Already had zip: {already_has}")
print(f"  Filled from FOIA: {filled}")
print(f"  License not found in FOIA: {not_found}")

# Write back
with open(VERIFIED, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

print(f"\nUpdated: {VERIFIED}")
