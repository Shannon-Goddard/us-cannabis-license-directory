import openpyxl, os

folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "csv")

for fname in sorted(os.listdir(folder)):
    if not fname.endswith(".xlsx"):
        continue
    wb = openpyxl.load_workbook(os.path.join(folder, fname), read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h) for h in (rows[0] if rows else []) if h is not None]
        print(f"{fname} - {len(rows)-1} rows")
        print(f"  {headers}")
    wb.close()
