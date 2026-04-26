"""
CO Facilities Builder
Written by Amazon Q for Loyal9 / poweredby.ci

Combines all Colorado MED Excel exports (same-schema files only) into:
  - CO-facilities.csv        — all records, all license types
  - CO-facilities-active.csv — non-expired records only

Source files collected: April 2026
Source: Colorado Marijuana Enforcement Division
        https://med.colorado.gov/licensee-information-and-lookup-tool
"""

import csv
import os
import openpyxl

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
STATE_DIR  = os.path.abspath(os.path.join(BASE_DIR, ".."))
CSV_DIR    = os.path.join(STATE_DIR, "csv")
ALL_OUT    = os.path.join(STATE_DIR, "CO-facilities.csv")
ACTIVE_OUT = os.path.join(STATE_DIR, "CO-facilities-active.csv")

# Only process files with the standard licensee schema
STANDARD_HEADERS = {
    "License Number", "Facility Name", "DBA",
    "Facility Type", "Street", "City", "ZIP Code",
    "Expiration Date", "Date Updated"
}

# Files to skip entirely
SKIP_FILES = {"Counts 2604 Apr.xlsx", "Testing Facilities 2604 Apr.xlsx"}

SEED_RELEVANT = {"Stores", "Retail Store", "Medical Store"}


def main():
    all_rows = []
    seen     = set()

    for fname in sorted(os.listdir(CSV_DIR)):
        if not fname.endswith(".xlsx") or fname in SKIP_FILES:
            continue

        fpath    = os.path.join(CSV_DIR, fname)
        category = fname.replace(" 2604 Apr.xlsx", "")
        wb       = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        file_count = 0

        for sheet in wb.sheetnames:
            ws      = wb[sheet]
            rows    = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(h).strip() if h else "" for h in rows[0]]

            # Skip sheets that don't match the standard schema
            if not STANDARD_HEADERS.issubset(set(headers)):
                continue

            idx = {h: i for i, h in enumerate(headers)}

            for row in rows[1:]:
                if not any(row):
                    continue
                lic_num = str(row[idx["License Number"]] or "").strip()
                if not lic_num or lic_num in seen:
                    continue
                seen.add(lic_num)

                all_rows.append({
                    "License_Number":  lic_num,
                    "Facility_Name":   str(row[idx["Facility Name"]] or "").strip(),
                    "DBA":             str(row[idx["DBA"]] or "").strip(),
                    "Facility_Type":   str(row[idx["Facility Type"]] or "").strip(),
                    "Street":          str(row[idx["Street"]] or "").strip(),
                    "City":            str(row[idx["City"]] or "").strip(),
                    "ZIP_Code":        str(row[idx["ZIP Code"]] or "").strip(),
                    "Expiration_Date": str(row[idx["Expiration Date"]] or "").strip(),
                    "Date_Updated":    str(row[idx["Date Updated"]] or "").strip(),
                    "Category":        category,
                    "Sheet":           sheet,
                })
                file_count += 1

        wb.close()
        print(f"  {fname}: {file_count} unique rows")

    HEADERS = [
        "License_Number", "Facility_Name", "DBA", "Facility_Type",
        "Street", "City", "ZIP_Code", "Expiration_Date",
        "Date_Updated", "Category", "Sheet"
    ]

    # Write master
    with open(ALL_OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=HEADERS)
        w.writeheader()
        w.writerows(all_rows)
    print(f"\nAll records:  {len(all_rows)} rows -> CO-facilities.csv")

    # Active = not expired (Expiration_Date is in the future or blank)
    from datetime import datetime
    today = datetime.today()

    def is_active(row):
        exp = row["Expiration_Date"]
        if not exp or exp in ("None", ""):
            return True
        try:
            return datetime.strptime(exp.split(" ")[0], "%Y-%m-%d") >= today
        except:
            try:
                return datetime.strptime(exp.split(" ")[0], "%m/%d/%Y") >= today
            except:
                return True  # keep if unparseable

    active = [r for r in all_rows if is_active(r)]
    with open(ACTIVE_OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=HEADERS)
        w.writeheader()
        w.writerows(active)
    print(f"Active only:  {len(active)} rows -> CO-facilities-active.csv")

    print("\nAll records by category:")
    cat_counts = {}
    for r in all_rows:
        c = r["Category"]
        cat_counts[c] = cat_counts.get(c, 0) + 1
    for c, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
        flag = " [seed finder]" if any(s in c for s in ["Store", "Cultivation"]) else ""
        print(f"  {n:>5}  {c}{flag}")


if __name__ == "__main__":
    main()
